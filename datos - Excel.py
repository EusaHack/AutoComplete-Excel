#Librerias 
from tkinter import *
import wmi
from openpyxl import Workbook
from openpyxl import load_workbook
import webbrowser


#Variables del Sistema para inf
c = wmi.WMI()
my_system = c.Win32_ComputerSystem()[0]
my_system1 = c.Win32_Bios()[0]
my_system2 = c.Win32_DiskDrive()[0]
my_system3 = c.Win32_PhysicalMemory()[0]

#Variables de Datos
usuarioRed = my_system.UserName
nombreEquipo = my_system.Name
marca = my_system.Manufacturer
modelo = my_system.SystemFamily
serie = my_system1.SerialNumber.strip()
discoDuro = my_system2.Size
ram = my_system3.Capacity

#Variable de interfaz
archivo = Tk()
#Configuracion de variable de Interfaz
archivo.title('Grupo')
archivo.geometry('600x350')
archivo.config(bg='white')

#Variables Texto
nombre = StringVar()
env = StringVar()


#Ingreso de Imagenes
img = PhotoImage(file='icono2.png')
Label(archivo, image= img,bg='white').pack()
img2 = PhotoImage(file='clavera2.png')
Label(archivo,image = img2,bg='white').place(x=320, y=320)

#Ingreso de Etiquetas
contacto = Label(archivo,text='By EusaHack',font='Curier 8', bg='white').place(x=250,y=320)
Label(archivo, text='Ingresa tu nombre :',bg='white',font='Curier 13').place(x=80,y=135)
mensaje = Label(archivo, textvariable=env, bg='white').place(x = 10, y=290 )
#Ingreso de Texto
txt1 = Entry(archivo,bd=3, textvariable = nombre ,bg ='gainsboro').place(x=250, y=135)

#Botones
enviar = Button(archivo,text='Enviar', command = lambda : insertarDatos()).place(x = 290, y = 180)
info = Button(archivo,text= '®', bg='white', command = lambda : abrirUrl()).place(x= 335, y =320)

#Funciones
def abrirUrl():
    url = webbrowser.open("https://www.linkedin.com/in/EusaHack/", new=2, autoraise=True)
def insertarDatos():
    
    #Cargar Archivo Excel
    wb = load_workbook('Doc-Excel.xlsx')

    #Ingresar a sheet
    ws = wb['Hoja1']
    #Ingresar datos 
    ws['B6'] = nombre.get()
    ws['H6'] = usuarioRed
    ws['H9'] = nombreEquipo
    ws['B17'] = marca
    ws['C17'] = modelo
    ws['D17'] = serie
    ws['F17'] = f'Disco Duro de {discoDuro[:3]} GB y {ram[:1]} GB en Memoria RAM '


    #Guardar Archivo Excel
    wb.save('Doc-Excel.xlsx')

    #Mensaje
    env.set('Guardado')

#Variable de interfaz
archivo.mainloop()



